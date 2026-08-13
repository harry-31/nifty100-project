import re
import sqlite3
from pathlib import Path

import pandas as pd

DB_PATH = Path(__file__).resolve().parents[2] / "nifty100.db"


def load_peer_data():
    """
    Load financial ratios joined with peer group information.
    """
    conn = sqlite3.connect(DB_PATH)

    query = """
    SELECT
        fr.*,
        pg.peer_group_name,
        pg.is_benchmark
    FROM financial_ratios fr
    JOIN peer_groups pg
        ON fr.company_id = pg.company_id
    """

    df = pd.read_sql_query(query, conn)
    conn.close()

    return df


def calculate_peer_rankings():
    """
    Calculate quality score and rank companies within each peer group.
    """
    df = load_peer_data().copy()

    # Quality score
    df["quality_score"] = (
        df["return_on_equity_pct"] * 0.30
        + df["operating_profit_margin_pct"] * 0.20
        + df["revenue_cagr"] * 0.20
        + df["pat_cagr"] * 0.20
        + df["interest_coverage"] * 0.10
    )

    # Helper to determine latest record
    def get_sort_year(value):
        value = str(value).strip()

        if value.upper() == "TTM":
            return 9999

        match = re.search(r"(19\d{2}|20\d{2})", value)
        if match:
            return int(match.group())

        return -1

    # Keep only latest record per company
    df["sort_year"] = df["year"].apply(get_sort_year)

    df = (
        df.sort_values("sort_year")
        .groupby("company_id", as_index=False)
        .last()
        .drop(columns="sort_year")
    )

    # Rank within peer group
    df["peer_rank"] = df.groupby("peer_group_name")["quality_score"].rank(
        method="dense", ascending=False
    )

    # Percentile
    df["peer_percentile"] = (
        df.groupby("peer_group_name")["quality_score"].rank(pct=True) * 100
    ).round(2)

    return df.sort_values(["peer_group_name", "peer_rank"])


def export_peer_comparison(filename="output/peer_comparison.xlsx"):
    """
    Export one Excel sheet per peer group with formatting.
    """
    from openpyxl.styles import PatternFill

    df = calculate_peer_rankings()

    output_path = Path(filename)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )

    yellow_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    gold_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        for group_name, group_df in df.groupby("peer_group_name"):

            group_df = group_df.sort_values("peer_rank").copy()

            # Add median row
            median_row = {}

            for column in group_df.columns:
                if pd.api.types.is_numeric_dtype(group_df[column]):
                    median_row[column] = group_df[column].median()
                else:
                    median_row[column] = "Median" if column == "company_id" else ""

            group_df = pd.concat(
                [group_df, pd.DataFrame([median_row])], ignore_index=True
            )

            sheet_name = str(group_name)[:31]

            group_df.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]

            headers = [cell.value for cell in worksheet[1]]

            # Highlight benchmark company
            if "is_benchmark" in headers:

                benchmark_col = headers.index("is_benchmark") + 1

                for row in range(2, worksheet.max_row + 1):

                    if worksheet.cell(row, benchmark_col).value == 1:

                        for col in range(1, worksheet.max_column + 1):
                            worksheet.cell(row, col).fill = gold_fill

            # Colour percentile column
            if "peer_percentile" in headers:

                percentile_col = headers.index("peer_percentile") + 1

                for row in range(2, worksheet.max_row + 1):

                    cell = worksheet.cell(row, percentile_col)

                    if cell.value is None:
                        continue

                    try:
                        value = float(cell.value)
                    except (ValueError, TypeError):
                        continue

                    if value >= 75:
                        cell.fill = green_fill
                    elif value >= 25:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill

    return output_path
